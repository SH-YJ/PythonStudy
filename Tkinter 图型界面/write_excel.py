# -- coding:UTF-8 --
import tkinter, openpyxl, os, threading
from openpyxl import workbook
import tkinter.messagebox as messagebox


def writetoexcel(x):  # 写入excel
    data = openpyxl.load_workbook('abc.xlsx')
    # 取第一张表
    wb = data.sheetnames
    table = data[wb[0]]
    print(table.title)  # 输出表名
    nrows = table.max_row  # 获得行数
    ncolumns = table.max_column  # 获得行数
    # i = old_rows
    table.cell(nrows + 1, int(entry4.get()), value=x)  # 写入数据
    # sheet1.write(i , 0, u)
    # sheet1.write(i , 1, p)
    # sheet1.write(i , 2, titles[var])
    data.save('abc.xlsx')


def sheet_method(work_book, add_sheet=None):  # 创建excel中的表
    if add_sheet is None:
        add_sheet = []
    wk = work_book
    ss_sheet = wk['Sheet']
    ss_sheet.title = add_sheet[0]
    for i in range(1, len(add_sheet)):
        wk.create_sheet(add_sheet[i])
    sheet_num = wk.sheetnames
    last_sheet = len(sheet_num) - 1
    sheet_index = sheet_num.index(sheet_num[last_sheet])
    wk.active = sheet_index


def create_excel(filename="filename.xlsx", recreate=False, add_sheet=None):  # 创建excel
    if add_sheet is None:
        add_sheet = []
    path = os.path.join(os.getcwd() + '\\' + filename + ".xlsx")
    p1 = os.path.exists(path)
    if p1:
        if recreate:
            os.remove(path)
            wk = workbook.Workbook()
            sheet_method(wk, add_sheet)
            wk.save(path)
            messagebox.showinfo('提示', '文件已存在')
        else:
            pass
    else:
        wk = workbook.Workbook()
        sheet_method(wk, add_sheet)
        wk.save(path)
        messagebox.showinfo('提示', '创建完成' + filename + '.xlsx')


def thread_it(func, *args):  # 开线程防止卡死
    t = threading.Thread(target=func, args=args)
    t.setDaemon(True)
    t.start()


# create_excel("123", recreate=True, add_sheet=["12","23"])

win = tkinter.Tk()
win.title("Excel")
win.geometry("400x400+200+50")

lable1 = tkinter.Label(win, text="文件名:", fg="pink", bg="grey", width=9, height=1, font=("黑体", 9), justify="left",
                       anchor="w")
lable1.pack()
entry1 = tkinter.Entry(win, width=50)
entry1.pack()

lable2 = tkinter.Label(win, text="表名:", fg="pink", bg="grey", width=9, height=1, font=("黑体", 9), justify="left",
                       anchor="w")
lable2.pack()
entry2 = tkinter.Entry(win, width=50)
entry2.pack()

lable3 = tkinter.Label(win, text="若不存在该表，请创建", fg="orange", width=20, height=1, font=("黑体", 9), justify="left",
                       anchor="w")
lable3.pack()
button1 = tkinter.Button(win, text="创建", bg="orange", fg="purple",
                         command=lambda: create_excel(entry1.get(), recreate=True, add_sheet=[entry2.get()])).pack()

lable4 = tkinter.Label(win, text="写入数据", fg="black", width=8, height=1, font=("楷体", 15), justify="left",
                       anchor="w").pack()

lable6 = tkinter.Label(win, text="列号", fg="black", width=8, height=1, font=("楷体", 8), justify="left", anchor="w").pack()
entry4 = tkinter.Entry(win, width=50)
entry4.pack()
lable7 = tkinter.Label(win, text="数据内容", fg="black", width=8, height=1, font=("楷体", 8), justify="left",
                       anchor="w").pack()
entry5 = tkinter.Entry(win, width=50)
entry5.pack()

button2 = tkinter.Button(win, text="写入", bg="orange", fg="purple", command=lambda: writetoexcel(entry5.get())).pack()

win.mainloop()
