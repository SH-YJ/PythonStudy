from xml.dom.minidom import parse
import openpyxl


def readxml():
    domTree = parse("web5.xml")
    # 文档根元素
    rootBode = domTree.documentElement
    print(rootBode.nodeName)

    webs = rootBode.getElementsByTagName("url")
    for web in webs:
        # albumurl 元素
        albumurl = web.getElementsByTagName("albumurl")[0]
        print(albumurl.nodeName, ":", albumurl.childNodes[0].data)
        # webname 元素
        webname = web.getElementsByTagName("webname")[0]
        print(webname.nodeName, ":", webname.childNodes[0].data)
        writetoexcel(albumurl.childNodes[0].data, webname.childNodes[0].data)


def writetoexcel(x, y):  # 写入excel
    data = openpyxl.load_workbook('xyz.xlsx')
    # 取第一张表
    wb = data.sheetnames
    table = data[wb[1]]  # 读取第二张表
    print(table.title)  # 输出表名
    nrows = table.max_row  # 获得行数
    ncolumns = table.max_column  # 获得行数
    table.cell(nrows + 1, 1, value=x)  # 写入数据
    table.cell(nrows + 1, 2, value=y)  # 写入数据
    data.save('xyz.xlsx')


if __name__ == '__main__':
    readxml()
