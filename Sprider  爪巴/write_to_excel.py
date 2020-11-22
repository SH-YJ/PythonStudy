# -- coding:UTF-8 --
import requests
from bs4 import BeautifulSoup
import os
import re
import openpyxl

'''
思路：获取网址
      获取图片地址
      爬取图片并保存
'''


# 获取网址
def getUrl(url):
    try:
        read = requests.get(url)  # 获取url
        read.raise_for_status()  # 状态响应 返回200连接成功
        read.encoding = read.apparent_encoding  # 从内容中分析出响应内容编码方式
        return read.text  # Http响应内容的字符串，即url对应的页面内容
    except:
        return "连接失败！"


if __name__ == '__main__':
    for li in range(13, 50):
        html_url = getUrl('https://www.192td.com/gq/xiuren/index_' + str(li) + '.html')
        soup = BeautifulSoup(html_url, "html.parser")
        data = openpyxl.load_workbook('abc.xlsx')
        # 取第一张表
        wb = data.sheetnames
        table = data[wb[1]]
        print(table.title)  # 输出表名
        nrows = table.max_row  # 获得行数
        ncolumns = table.max_column  # 获得行数
        urls = soup.find('ul', class_='clearfix').find_all('a')
        titles = re.findall('<span>(.*?)</span>', html_url)
        # i = old_rows
        var = 1
        for url in urls:
            u = url['href']
            if len(u) < 30:
                u = 'https://www.192td.com' + u
            else:
                u = u
            u_html_url = getUrl(u)
            soups = BeautifulSoup(u_html_url, "html.parser")
            page = soups.find('h1').find('span', id='allnum')
            p = page.string
            table.cell(nrows + 1, 1, value=u)  # 写入数据
            table.cell(nrows + 1, 2, value=p)
            table.cell(nrows + 1, 3, value=titles[var])
            nrows = nrows + 1
            # sheet1.write(i , 0, u)
            # sheet1.write(i , 1, p)
            # sheet1.write(i , 2, titles[var])
            var = var + 1
        data.save('abc.xlsx')
