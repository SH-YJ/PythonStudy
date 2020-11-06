import requests
from bs4 import BeautifulSoup
import re
import math
import os
import openpyxl

headers ={
    'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/78.0.3904.116 Safari/537.36'
}

def getUrl(url):
    try:
        read = requests.get(url,headers=headers)  #获取url
        read.raise_for_status()   #状态响应 返回200连接成功
        read.encoding = read.apparent_encoding  #从内容中分析出响应内容编码方式
        return read.text   #Http响应内容的字符串，即url对应的页面内容
    except :
        return

def writetoexcel(z):#写入excel
    data = openpyxl.load_workbook('xyz.xlsx')
    # 取第一张表
    wb = data.sheetnames
    table = data[wb[0]]
    print(table.title)  # 输出表名
    nrows = table.max_row  # 获得行数
    ncolumns = table.max_column  # 获得列数
    # i = old_rows
    #table.cell(nrows + 1, 1, value=x)  # 写入数据
    #able.cell(nrows + 1, 2, value=y)
    table.cell(nrows + 1, 4, value=z)
    # sheet1.write(i , 0, u)
    # sheet1.write(i , 1, p)
    # sheet1.write(i , 2, titles[var])
    data.save('xyz.xlsx')

if __name__ == '__main__':
    html = getUrl("https://2flj.xyz/ziyuans/wp/page/4")
    soup = BeautifulSoup(html, "html.parser")
    urls = soup.find('div', class_='row').find_all('a', class_='media-content')
    for url in urls:
        i = url['href']
        html2 = getUrl(i)
        soup2 = BeautifulSoup(html2,"html.parser")
        dizhi = re.findall('<div class="col-6 col-md-7">(.*?)</div>',html2)
        code = re.findall('<div class="col-2 col-md-2" style=".*?">(.*?)</div>',html2)
        name = re.findall('<div class="site-name h3 my-3">(.*?)</div>',html2)
        jieya = re.findall('<span style="color: #ff0000;">(.*?)</span>',html2)
        writetoexcel(jieya[0])
