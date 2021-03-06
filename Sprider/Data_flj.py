import requests
from bs4 import BeautifulSoup
import re
import math
import os
import openpyxl

headers = {
    'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/78.0.3904.116 Safari/537.36'
}


def getUrl(url):
    try:
        read = requests.get(url, headers=headers)  # 获取url
        read.raise_for_status()  # 状态响应 返回200连接成功
        read.encoding = read.apparent_encoding  # 从内容中分析出响应内容编码方式
        return read.text  # Http响应内容的字符串，即url对应的页面内容
    except:
        return


def writetoexcel(x, y, z):  # 写入excel
    data = openpyxl.load_workbook('flj.xlsx')
    # 取第一张表
    wb = data.sheetnames
    table = data[wb[0]]
    print(table.title)  # 输出表名
    nrows = table.max_row  # 获得行数
    ncolumns = table.max_column  # 获得列数
    # i = old_rows
    table.cell(nrows + 1, 1, value=x)  # 写入数据
    table.cell(nrows + 1, 2, value=y)  # 写入数据
    table.cell(nrows + 1, 3, value=z)  # 写入数据
    # table.cell(nrows + 1, 4, value=a)  # 写入数据
    # table.cell(nrows + 1, 2, value=y)
    # table.cell(nrows + 1, 3, value=z)
    data.save('flj.xlsx')


if __name__ == '__main__':
    html = getUrl("https://2flj.xyz/ziyuans/wp")
    soup = BeautifulSoup(html, "html.parser")
    urls = soup.find('div', class_='row').find_all('a', class_='media-content')
    for url in urls:
        i = url['href']
        html2 = getUrl(i)
        soup2 = BeautifulSoup(html2, "html.parser")
        dizhi1 = re.findall('<div class="col-6 col-md-7">(.*?)</div>', html2)  # 百度网盘地址
        dizhi2 = re.findall('<a class="btn btn-danger custom_btn-d py-0 px-1 mx-auto down_count text-sm" href="(.*?)" target="_blank" data-id=".*?" data-action="down_count" data-mmid="down-mm-1">.*?</a>', html2)  # 黑料姬网盘地址
        code = re.findall('<div class="col-2 col-md-2" style=".*?">(.*?)</div>', html2)  # 提取码
        name = re.findall('<div class="site-name h3 my-3">(.*?)</div>', html2)  # 文件名
        jieya = re.findall('<span style="color: #ff0000;">(.*?)</span>', html2)  # 解压回复数字
        print(dizhi1[1], name[0], jieya[0])