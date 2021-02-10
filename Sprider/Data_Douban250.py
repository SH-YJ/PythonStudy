import urllib.error
import urllib.request
import re
from bs4 import BeautifulSoup
import time
import openpyxl

header = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/80.0.3987.163 Safari/537.36'
}
baseurl = "https://movie.douban.com/top250?start="


def getUrl(url):
    request = urllib.request.Request(url, headers=header)
    try:
        response = urllib.request.urlopen(request)
        html = response.read().decode("utf-8")
        return html
    except urllib.error.URLError as e:
        if hasattr(e, 'code'):
            return e.code
        if hasattr(e, 'reason'):
            return e.reason


if __name__ == '__main__':
    for x in range(0, 10):
        html = getUrl(baseurl + str(x*25))
        Ch_movie_name = re.findall('<span class="title">([\u4E00-\u9FA5]+)</span>', html)  # 获得中文名
        Score = re.findall('<span class="rating_num" property="v:average">(.*?)</span>', html)  # 获取评分
        soup = BeautifulSoup(html, "html.parser")
        M_url = soup.find('ol', class_='grid_view').find_all('a', class_='')
        time.sleep(1)
        li = []
        for i in range(0, len(M_url), 2):
            mm_url = M_url[i]['href']
            M_html = getUrl(mm_url)
            Director = re.findall('<a href=".*?" rel="v:directedBy">(.*?)</a>', M_html)
            li.append(Director)
            time.sleep(1)
        data = openpyxl.load_workbook('douban.xlsx')
        # 取第一张表
        wb = data.sheetnames
        table = data[wb[0]]
        print(table.title)  # 输出表名
        nrows = table.max_row  # 获得行数
        ncolumns = table.max_column  # 获得行数
        for i, j, k in zip(Ch_movie_name, li, Score):
            table.cell(nrows+1, 1, value=i)
            table.cell(nrows+1, 2, value=j[0])
            table.cell(nrows+1, 3, value=k)
            nrows = nrows + 1
        data.save('douban.xlsx')
        time.sleep(1)
